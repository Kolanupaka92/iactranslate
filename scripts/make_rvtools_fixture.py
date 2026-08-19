"""Generate a *structurally realistic* RVTools workbook.

The committed `rvtools_sample.xlsx` fixture has 3 sheets and 10 columns. A real
RVTools export has 11+ sheets and ~40 columns on `vInfo` alone, and Azure
Migrate documents files of up to 20,000 servers. Testing only against the tidy
fixture means the parser has never seen the shape it will actually be given.

Sheet names and the column sets Azure Migrate reads are taken from Microsoft's
RVTools import spec:
https://learn.microsoft.com/en-us/azure/migrate/tutorial-import-vmware-using-rvtools-xlsx

Usage:
    python scripts/make_rvtools_fixture.py                      # 25-VM test fixture
    python scripts/make_rvtools_fixture.py --vms 5000 --out /tmp/big.xlsx
"""
from __future__ import annotations

import argparse
import random
from pathlib import Path

from openpyxl import Workbook

# RVTools reports guest OS with these exact strings.
GUEST_OS = [
    "Ubuntu Linux (64-bit)",
    "Red Hat Enterprise Linux 8 (64-bit)",
    "Microsoft Windows Server 2019 (64-bit)",
    "Microsoft Windows Server 2016 (64-bit)",
    "CentOS 7 (64-bit)",
    "Microsoft Windows Server 2012 R2 (64-bit)",
    "SUSE Linux Enterprise 15 (64-bit)",
    "Other 4.x or later Linux (64-bit)",
]
TIERS = ["web", "app", "db", "cache", "mq", "batch"]
ENVS = ["prod", "stg", "dev", "uat"]

# The real vInfo sheet is wide. This is a representative subset in RVTools'
# own column order, including the eight Azure Migrate requires.
VINFO_COLUMNS = [
    "VM", "Powerstate", "Template", "SRM Placeholder", "Config status", "DNS Name",
    "Connection state", "Guest state", "Heartbeat", "Consolidation Needed",
    "PowerOn", "Suspend time", "Creation date", "Change Version",
    "CPUs", "Latency Sensitivity", "Memory", "NICs", "Disks",
    "Total disk capacity MiB", "min Required EVC Mode Key",
    "Provisioned MiB", "In Use MiB", "Unshared MiB",
    "HA Restart Priority", "HA Isolation Response", "Cluster rule(s)",
    "Boot Required", "Boot delay", "Boot retry delay", "Boot BIOS setup",
    "Firmware", "HW version", "HW upgrade status", "HW upgrade policy",
    "Path", "Log directory", "Snapshot directory", "Suspend directory",
    "Annotation", "Datacenter", "Cluster", "Host", "OS according to the configuration file",
    "OS according to the VMware Tools", "VM ID", "VM UUID", "VI SDK Server", "VI SDK UUID",
]


def _vm_names(count: int) -> list:
    """Names shaped like a real estate: mixed case, separators, and spaces."""
    names, i = [], 0
    while len(names) < count:
        env, tier = ENVS[i % len(ENVS)], TIERS[i % len(TIERS)]
        n = i // (len(ENVS) * len(TIERS)) + 1
        # Real inventories are inconsistent; rotate through plausible styles.
        style = i % 5
        if style == 0:
            names.append(f"{env}-{tier}-{n:02d}")
        elif style == 1:
            names.append(f"{env.upper()}_{tier.upper()}_{n:02d}")
        elif style == 2:
            names.append(f"{env}.{tier}.{n:02d}.corp.local")
        elif style == 3:
            names.append(f"{env} {tier} {n:02d}")          # spaces happen
        else:
            names.append(f"{tier}{n:03d}")
        i += 1
    return names


def build(vms: int, out: Path, seed: int = 11) -> Path:
    rng = random.Random(seed)
    wb = Workbook(write_only=True)
    names = _vm_names(vms)

    vinfo = wb.create_sheet("vInfo")
    vinfo.append(VINFO_COLUMNS)
    disks_per_vm, nics_per_vm, meta = {}, {}, {}

    for idx, name in enumerate(names):
        uuid = f"421a{idx:04d}-0000-0000-0000-{idx:012d}"
        powered = rng.random() > 0.12          # ~12% powered off, as in real estates
        cpus = rng.choice([1, 2, 2, 4, 4, 8, 8, 16, 32])
        mem_mib = rng.choice([2048, 4096, 8192, 16384, 32768, 65536, 131072])
        ndisks = rng.choice([1, 1, 1, 2, 2, 3, 5])
        disk_sizes = [rng.choice([40960, 51200, 102400, 204800, 512000]) for _ in range(ndisks)]
        provisioned = sum(disk_sizes)
        in_use = int(provisioned * rng.uniform(0.25, 0.9))
        nnics = rng.choice([1, 1, 1, 2])
        os_cfg = rng.choice(GUEST_OS)
        cluster = f"cluster-{idx % 8:02d}"
        dc = f"dc-{idx % 3:02d}"
        host = f"esx{idx % 40:02d}.corp.local"

        disks_per_vm[name] = disk_sizes
        nics_per_vm[name] = nnics
        meta[name] = (uuid, powered, cluster, dc, host, mem_mib)

        row = {
            "VM": name,
            "Powerstate": "poweredOn" if powered else "poweredOff",
            "Template": False,
            "Config status": "green",
            "DNS Name": f"{name.replace(' ', '-').replace('_', '-').lower()}",
            "Connection state": "connected",
            "Guest state": "running" if powered else "notRunning",
            "Heartbeat": "green" if powered else "gray",
            "Consolidation Needed": False,
            "CPUs": cpus,
            "Memory": mem_mib,
            "NICs": nnics,
            "Disks": ndisks,
            "Total disk capacity MiB": provisioned,
            "Provisioned MiB": provisioned,
            "In Use MiB": in_use,
            "Unshared MiB": in_use,
            "Firmware": rng.choice(["bios", "efi"]),
            "HW version": f"vmx-{rng.choice([13, 14, 15, 17, 19])}",
            "Path": f"[datastore{idx % 12}] {name}/{name}.vmx",
            "Datacenter": dc,
            "Cluster": cluster,
            "Host": host,
            "OS according to the configuration file": os_cfg,
            # Real files often have this blank when VMware Tools isn't running.
            "OS according to the VMware Tools": os_cfg if powered else None,
            "VM ID": f"vm-{1000 + idx}",
            "VM UUID": uuid,
            "VI SDK Server": "vcenter.corp.local",
            "VI SDK UUID": "aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee",
        }
        vinfo.append([row.get(col) for col in VINFO_COLUMNS])

    # vDisk: one row per virtual disk.
    vdisk = wb.create_sheet("vDisk")
    vdisk.append(["VM", "Powerstate", "Template", "Disk", "Capacity MiB", "Raw",
                  "Disk Mode", "Thin", "Shared Bus", "Controller", "Path",
                  "Datacenter", "Cluster", "Host", "VM UUID"])
    for name, sizes in disks_per_vm.items():
        uuid, powered, cluster, dc, host, _ = meta[name]
        for d, size in enumerate(sizes):
            vdisk.append([name, "poweredOn" if powered else "poweredOff", False,
                          f"Hard disk {d + 1}", size, False, "persistent",
                          bool(d % 2), "noSharing", "SCSI controller 0",
                          f"[datastore] {name}/{name}_{d}.vmdk", dc, cluster, host, uuid])

    # vNetwork: one row per NIC.
    vnet = wb.create_sheet("vNetwork")
    vnet.append(["VM", "Powerstate", "Template", "Adapter", "Network", "Switch",
                 "Connected", "Starts Connected", "Mac Address", "IP Address",
                 "Datacenter", "Cluster", "Host", "VM UUID"])
    for i, (name, nnics) in enumerate(nics_per_vm.items()):
        uuid, powered, cluster, dc, host, _ = meta[name]
        for n in range(nnics):
            vnet.append([name, "poweredOn" if powered else "poweredOff", False,
                         f"Network adapter {n + 1}", f"VLAN_{100 + (i % 12)}",
                         "dvSwitch01", powered, True,
                         f"00:50:56:{i % 256:02x}:{n:02x}:01",
                         f"10.{i // 254 % 254}.{i % 254}.{n + 10}" if powered else None,
                         dc, cluster, host, uuid])

    # vPartition: guest-visible partitions (real consumption).
    vpart = wb.create_sheet("vPartition")
    vpart.append(["VM", "Disk", "Capacity MiB", "Consumed MiB", "Free MiB",
                  "Free %", "Datacenter", "Cluster", "Host", "VM UUID"])
    for name, sizes in disks_per_vm.items():
        uuid, _, cluster, dc, host, _ = meta[name]
        for d, size in enumerate(sizes):
            consumed = int(size * rng.uniform(0.2, 0.85))
            vpart.append([name, "C:\\" if d == 0 else f"/data{d}", size, consumed,
                          size - consumed, round((size - consumed) / size * 100, 1),
                          dc, cluster, host, uuid])

    # vMemory
    vmem = wb.create_sheet("vMemory")
    vmem.append(["VM", "Size MiB", "Overhead", "Reservation", "Limit", "Shares",
                 "Ballooned", "Swapped", "Datacenter", "Cluster", "Host", "VM UUID"])
    for name in names:
        uuid, _, cluster, dc, host, mem_mib = meta[name]
        vmem.append([name, mem_mib, 96, 0, -1, "normal", 0, 0, dc, cluster, host, uuid])

    # Sheets we never read, but which exist in every real export — their
    # presence is the point: the parser must ignore them without loading cost.
    for sheet, cols in [
        ("vCPU", ["VM", "CPUs", "Sockets", "Cores p/s", "Reservation", "Limit", "VM UUID"]),
        ("vHost", ["Host", "Datacenter", "Cluster", "CPU Model", "Speed", "# CPU",
                   "Cores per CPU", "# Cores", "CPU usage %", "# Memory",
                   "Memory usage %", "ESX Version", "Vendor", "Model", "UUID"]),
        ("vDatastore", ["Name", "Object ID", "Type", "Hosts", "Capacity MiB",
                        "Provisioned MiB", "In Use MiB", "Free MiB", "Free %"]),
        ("vSnapshot", ["VM", "VM UUID", "Powerstate", "Name", "Date / time",
                       "Size MiB (vmsn)", "Size MiB (total)", "Quiesced",
                       "Datacenter", "Cluster", "Host"]),
        ("vCD", ["VM", "VM UUID", "Powerstate", "Device Type", "Connected"]),
        ("vUSB", ["VM", "VM UUID", "Powerstate", "Device Type", "Connected"]),
        ("vTools", ["VM", "Tools", "Tools Version", "Upgradeable", "VM UUID"]),
        ("dvPort", ["Object ID", "Port", "Switch", "Type", "VLAN",
                    "Allow Promiscuous", "Mac changes", "Forged Transmits"]),
        ("vMetaData", ["Data", "Value"]),
    ]:
        ws = wb.create_sheet(sheet)
        ws.append(cols)
        for i in range(min(len(names), 40)):
            ws.append([names[i] if c == "VM" else None for c in cols])

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--vms", type=int, default=25)
    ap.add_argument("--out", default="tests/fixtures/rvtools_realistic.xlsx")
    args = ap.parse_args()
    path = build(args.vms, Path(args.out))
    size_kb = path.stat().st_size / 1024
    print(f"wrote {path} — {args.vms} VMs, {size_kb:,.0f} KB")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
